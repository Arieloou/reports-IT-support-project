import flet as ft
from fpdf import FPDF
from openpyxl import Workbook, load_workbook
import os
import platform
import subprocess

# ------------------ LIMPIEZA ------------------
def clean_up_text(texto):
    reemplazos = {
        "á":"a","é":"e","í":"i","ó":"o","ú":"u",
        "Á":"A","É":"E","Í":"I","Ó":"O","Ú":"U",
        "ñ":"n","Ñ":"N",
        "–":"-","—":"-","“":'"',"”":'"'
    }
    for k, v in reemplazos.items():
        texto = texto.replace(k, v)
    return texto

# ------------------ BUSCAR EN EXCEL ------------------
def standardize_name(nombre):
    """Normaliza un nombre: limpia, pasa a minúsculas y ordena las palabras."""
    palabras = clean_up_text(nombre).strip().lower().split()
    palabras.sort()
    return palabras

## Esta funcion permite buscar el número de cédula asociado a una persona tomando en cuenta los nombres que han sido registrados
def get_identification_card(nombre_buscar):
    archivo = "DATA.xlsx"
    if not os.path.exists(archivo):
        return None

    wb = load_workbook(archivo)
    if "PERSONAS" not in wb.sheetnames:
        return None
    ws = wb["PERSONAS"]

    palabras_buscar = standardize_name(nombre_buscar)
    if not palabras_buscar:
        return None

    ## Si se ingresaron menos de 3 palabras, quiere decir que no colocaron los dos nombres o los dos apellidos
    possible_matches = []

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0] is None:
            continue
        palabras_fila = standardize_name(str(fila[0]))

        if set(palabras_buscar).issubset(palabras_fila):
            possible_matches.append((str(fila[0]), str(fila[1])))
    
    return possible_matches # Multiples coincidencias

    return None

# ------------------ ABRIR PDF ------------------
def open_pdf(ruta):
    sistema = platform.system()
    if sistema == "Windows":
        os.startfile(ruta)
    elif sistema == "Darwin":
        subprocess.call(["open", ruta])
    else:
        subprocess.call(["xdg-open", ruta])

# ------------------ PDF ------------------
def generate_pdf(data):
    class ActaPDF(FPDF):
        def header(self):
            self.set_font("Arial", "B", 12)
            self.cell(0, 10, "ACTA ENTREGA - RECEPCION", 0, 1, "C")

    pdf = ActaPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.multi_cell(0, 8,
    f"Yo {data['nombre']} portador del documento de identidad {data['cedula']}, declaro que he recibido...")

    pdf.ln(5)

    def fila(t, v):
        pdf.cell(60, 8, t, 1)
        pdf.cell(130, 8, v, 1)
        pdf.ln()

    fila("Equipo(s)", data["equipo"])
    fila("Marca(s)", data["marca"])
    fila("Modelo(s)", data["modelo"])
    fila("Codigo", data["codigo"])
    fila("Ticket", data["ticket"])
    fila("Fecha", data["fecha"])
    fila("Observaciones", data["observaciones"])

    pdf.output("Acta_Entrega.pdf")
    return "Acta_Entrega.pdf"

# ------------------ EXCEL ------------------
def save_record_in_excel(data):
    archivo = "DATA.xlsx"

    if os.path.exists(archivo):
        wb = load_workbook(archivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nombre","Cedula","Fecha","Equipo","Marca","Modelo","Codigo","Ticket","Observaciones"])

    ws.append(list(data.values()))
    wb.save(archivo)

# ------------------ APP ------------------
def main(page: ft.Page):
    page.title = "Sistema de Actas"
    page.window_width = 500

    # ---------- FORMULARIO ----------
    def form_view():
        nombre = ft.TextField(label="Nombres")
        cedula = ft.TextField(label="Cedula")
        fecha = ft.TextField(label="Fecha")

        equipo = ft.TextField(label="Equipo")
        marca = ft.TextField(label="Marca")
        modelo = ft.TextField(label="Modelo")
        codigo = ft.TextField(label="Codigo")
        ticket = ft.TextField(label="Ticket")
        observaciones = ft.TextField(label="Observaciones")

        def show_popup_message(texto):
            def close_popup(e):
                dialogo.open = False
                page.update()
            dialogo = ft.AlertDialog(
                title=ft.Text("Estado"),
                content=ft.Text(texto),
                actions=[ft.TextButton("Aceptar", on_click=close_popup)],
            )
            page.overlay.append(dialogo)
            dialogo.open = True
            page.update()

        def show_individual_options_list(matches):
            """Muestra un dialogo con las coincidencias para que el usuario elija una."""
            def seleccionar(e, cedula_sel, nombre_sel):
                cedula.value = cedula_sel
                nombre.value = nombre_sel
                dialogo.open = False
                page.update()

            opciones = []
            for nom, ced in matches:
                opciones.append(
                    ft.TextButton(
                        content=ft.Text(f"{nom}  —  {ced}"),
                        on_click=lambda e, c=ced, n=nom: seleccionar(e, c, n),
                    )
                )

            def close_options_list(e):
                dialogo.open = False
                page.update()

            dialogo = ft.AlertDialog(
                title=ft.Text("Seleccione una persona"),
                content=ft.Column(opciones, tight=True, scroll=ft.ScrollMode.AUTO),
                actions=[ft.TextButton("Cancelar", on_click=close_options_list)],
            )
            page.overlay.append(dialogo)
            dialogo.open = True
            page.update()

        def search_individual_info(e):
            if not nombre.value or not nombre.value.strip():
                show_popup_message("Ingrese un nombre para realizar la búsqueda.")
                return

            personal_info = get_identification_card(nombre.value)

            if not personal_info:
                show_popup_message("No se encontro información para la persona ingresada.")
            else:
                show_individual_options_list(personal_info)

        search_button = ft.IconButton(
            icon=ft.Icons.SEARCH,
            tooltip="Buscar cedula",
            on_click=search_individual_info,
        )

        fila_nombre = ft.Row([nombre, search_button])

        def obtener():
            return {
                "nombre": clean_up_text(nombre.value),
                "cedula": clean_up_text(cedula.value),
                "fecha": clean_up_text(fecha.value),
                "equipo": clean_up_text(equipo.value),
                "marca": clean_up_text(marca.value),
                "modelo": clean_up_text(modelo.value),
                "codigo": clean_up_text(codigo.value),
                "ticket": clean_up_text(ticket.value),
                "observaciones": clean_up_text(observaciones.value)
            }

        def generar(e):
            ruta = generate_pdf(obtener())
            open_pdf(ruta)
            show_popup_message("PDF generado")

        def guardar(e):
            save_record_in_excel(obtener())
            show_popup_message("Guardado en Excel")

        def volver(e):
            menu()

        page.controls.clear()
        page.add(
            ft.Column([
                ft.Text("Acta de Entrega", size=20),
                fila_nombre, cedula, fecha,
                equipo, marca, modelo, codigo, ticket, observaciones,
                ft.Row([
                    ft.ElevatedButton("PDF", on_click=generar),
                    ft.ElevatedButton("Excel", on_click=guardar),
                    ft.ElevatedButton("Volver", on_click=volver),
                ])
            ])
        )

    # ---------- MENÚ ----------
    def menu():
        page.controls.clear()
        page.add(
            ft.Column([
                ft.Text("MENU PRINCIPAL", size=25, weight="bold"),

                ft.ElevatedButton("📄 Generar Acta de Entrega", on_click=lambda e: form_view()),
                ft.ElevatedButton("🔄 Acta de Devolucion", on_click=lambda e: form_view()),
                ft.ElevatedButton("🎧 Accesorios Complementarios", on_click=lambda e: form_view()),
                ft.ElevatedButton("❌ Salir", on_click=lambda e: page.window_close())
            ])
        )

    # iniciar menú
    menu()

# RUN
ft.app(target=main)