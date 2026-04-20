import flet as ft
from fpdf import FPDF
from openpyxl import Workbook, load_workbook
import os
import platform
import subprocess

# ------------------ LIMPIEZA ------------------
def limpiar(texto):
    reemplazos = {
        "á":"a","é":"e","í":"i","ó":"o","ú":"u",
        "Á":"A","É":"E","Í":"I","Ó":"O","Ú":"U",
        "ñ":"n","Ñ":"N",
        "–":"-","—":"-","“":'"',"”":'"'
    }
    for k, v in reemplazos.items():
        texto = texto.replace(k, v)
    return texto

# ------------------ BUSCAR EN EXCEL ------------------
def buscar_cedula(nombre_buscar):
    archivo = "registros.xlsx"
    if not os.path.exists(archivo):
        return None

    wb = load_workbook(archivo)
    ws = wb.active

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if str(fila[0]).strip().lower() == nombre_buscar.strip().lower():
            return str(fila[1])
    return None

# ------------------ ABRIR PDF ------------------
def abrir_pdf(ruta):
    sistema = platform.system()
    if sistema == "Windows":
        os.startfile(ruta)
    elif sistema == "Darwin":
        subprocess.call(["open", ruta])
    else:
        subprocess.call(["xdg-open", ruta])

# ------------------ PDF ------------------
def generar_pdf(data):
    class ActaPDF(FPDF):
        def header(self):
            self.set_font("Arial", "B", 12)
            self.cell(0, 10, "ACTA ENTREGA - RECEPCION", 0, 1, "C")

    pdf = ActaPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.multi_cell(0, 8,
    f"Yo {data['nombre']} portador del documento de identidad {data['cedula']}, declaro que he recibido...")

    pdf.ln(5)

    def fila(t, v):
        pdf.cell(60, 8, t, 1)
        pdf.cell(130, 8, v, 1)
        pdf.ln()

    fila("Equipo(s)", data["equipo"])
    fila("Marca(s)", data["marca"])
    fila("Modelo(s)", data["modelo"])
    fila("Codigo", data["codigo"])
    fila("Ticket", data["ticket"])
    fila("Fecha", data["fecha"])
    fila("Observaciones", data["observaciones"])

    pdf.output("Acta_Entrega.pdf")
    return "Acta_Entrega.pdf"

# ------------------ EXCEL ------------------
def guardar_excel(data):
    archivo = "registros.xlsx"

    if os.path.exists(archivo):
        wb = load_workbook(archivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nombre","Cedula","Fecha","Equipo","Marca","Modelo","Codigo","Ticket","Observaciones"])

    ws.append(list(data.values()))
    wb.save(archivo)

# ------------------ APP ------------------
def main(page: ft.Page):
    page.title = "Sistema de Actas"
    page.window_width = 500

    # ---------- FORMULARIO ----------
    def vista_formulario():
        nombre = ft.TextField(label="Nombres")
        cedula = ft.TextField(label="Cedula")
        fecha = ft.TextField(label="Fecha")

        equipo = ft.TextField(label="Equipo")
        marca = ft.TextField(label="Marca")
        modelo = ft.TextField(label="Modelo")
        codigo = ft.TextField(label="Codigo")
        ticket = ft.TextField(label="Ticket")
        observaciones = ft.TextField(label="Observaciones")

        mensaje = ft.Text()

        def on_nombre_change(e):
            c = buscar_cedula(limpiar(nombre.value))
            if c:
                cedula.value = c
                page.update()

        nombre.on_change = on_nombre_change

        def obtener():
            return {
                "nombre": limpiar(nombre.value),
                "cedula": limpiar(cedula.value),
                "fecha": limpiar(fecha.value),
                "equipo": limpiar(equipo.value),
                "marca": limpiar(marca.value),
                "modelo": limpiar(modelo.value),
                "codigo": limpiar(codigo.value),
                "ticket": limpiar(ticket.value),
                "observaciones": limpiar(observaciones.value)
            }

        def generar(e):
            ruta = generar_pdf(obtener())
            abrir_pdf(ruta)
            mensaje.value = "PDF generado"
            page.update()

        def guardar(e):
            guardar_excel(obtener())
            mensaje.value = "Guardado en Excel"
            page.update()

        def volver(e):
            menu()

        page.controls.clear()
        page.add(
            ft.Column([
                ft.Text("Acta de Entrega", size=20),
                nombre, cedula, fecha,
                equipo, marca, modelo, codigo, ticket, observaciones,
                ft.Row([
                    ft.ElevatedButton("PDF", on_click=generar),
                    ft.ElevatedButton("Excel", on_click=guardar),
                    ft.ElevatedButton("Volver", on_click=volver),
                ]),
                mensaje
            ])
        )

    # ---------- MENÚ ----------
    def menu():
        page.controls.clear()
        page.add(
            ft.Column([
                ft.Text("MENU PRINCIPAL", size=25, weight="bold"),

                ft.ElevatedButton("📄 Generar Acta de Entrega", on_click=lambda e: vista_formulario()),
                ft.ElevatedButton("🔄 Acta de Devolucion", on_click=lambda e: vista_formulario()),
                ft.ElevatedButton("🎧 Accesorios Complementarios", on_click=lambda e: vista_formulario()),
                ft.ElevatedButton("❌ Salir", on_click=lambda e: page.window_close())
            ])
        )

    # iniciar menú
    menu()

# RUN
ft.app(target=main)