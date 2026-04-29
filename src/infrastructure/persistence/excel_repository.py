import os
from openpyxl import Workbook, load_workbook
from src.core.utils import standardize_name
from src.core.constants import DATA_FILE_PATH
from src.domain.models import Person, NewPersonRecord, TransactionRecord
from typing import List

class ExcelRepository:
    def __init__(self, file_path: str = DATA_FILE_PATH):
        self.file_path = file_path

    def get_personas_by_name(self, nombre_buscar: str) -> List[Person]:
        """Busca personas por nombre en el archivo de Excel."""
        if not os.path.exists(self.file_path):
            return []

        try:
            wb = load_workbook(self.file_path, read_only=True)
            if "PERSONAS" not in wb.sheetnames:
                wb.close()
                return []
            
            ws = wb["PERSONAS"]
            palabras_buscar = standardize_name(nombre_buscar)
            
            if not palabras_buscar:
                wb.close()
                return []

            possible_matches = []
            for fila in ws.iter_rows(min_row=2, values_only=True):
                if fila[0] is None:
                    continue
                palabras_fila = standardize_name(str(fila[0]))

                if set(palabras_buscar).issubset(palabras_fila):
                    possible_matches.append(Person(nombre=str(fila[0]), cedula=str(fila[1])))
            
            wb.close()
            return possible_matches
        except Exception as e:
            print(f"Error reading Excel for search: {e}")
            return []

    def save_acta_record(self, data: TransactionRecord) -> bool:
        """Guarda el registro de un acta generada en el Excel."""
        try:
            if os.path.exists(self.file_path):
                wb = load_workbook(self.file_path)
                ws = wb["REGISTROS"]
            else:
                wb = Workbook()
                ws = wb["REGISTROS"]
                ws.append([
                    "NOMBRE", "CEDULA", "Fecha",
                    "Equipo", "Marca", "Modelo", "Codigo",
                    "Equipo 2", "Marca 2", "Modelo 2", "Codigo 2",
                    "Ticket", "Observaciones"
                ])

            ws.append([
                data.nombre,
                data.cedula,
                data.fecha,
                data.equipo,
                data.marca,
                data.modelo,
                data.codigo,
                data.equipo2,
                data.marca2,
                data.modelo2,
                data.codigo2,
                data.ticket,
                data.observaciones,
            ])
            wb.save(self.file_path)
            return True
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            return False

    def save_acta_record(self, data: NewPersonRecord) -> bool:
        """Guarda el registro de un acta generada en el Excel."""
        try:
            if os.path.exists(self.file_path):
                wb = load_workbook(self.file_path)
                ws = wb["REGISTROS"]
            else:
                wb = Workbook()
                ws = wb["REGISTROS"]
                ws.append([
                    "NOMBRE", "CEDULA", "Fecha",
                    "Equipo", "Marca", "Modelo", "Codigo",
                    "Equipo 2", "Marca 2", "Modelo 2", "Codigo 2",
                    "Ticket", "Observaciones"
                ])

            ws.append([
                data.nombre,
                data.cedula,
                data.fecha,
                data.equipo,
                data.marca,
                data.modelo,
                data.codigo,
                data.equipo2,
                data.marca2,
                data.modelo2,
                data.codigo2,
                data.ticket,
                data.observaciones,
            ])
            wb.save(self.file_path)
            return True
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            return False
